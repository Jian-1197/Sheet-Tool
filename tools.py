import re
import os
import sys
import zipfile
import shutil
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side

from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT, TA_RIGHT

# 中文名字筛选
def is_chinese(name):
    return bool(re.search(r'^[\u4e00-\u9fa5]+$', name))

# 周/月考勤表表格数据写入及格式设置
def write_attendance_sheet(ws, data, columns, col_widths, hidden_cols):

    # 样式设置变量
    表头字体 = Font(size=11, bold=True)     # 表头字体样式
    内容字体 = Font(size=11)                # 内容字体样式
    边框样式 = Border(                      # 单元格边框样式    
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    对齐方式 = Alignment(horizontal="center", vertical="center")  # 居中对齐
    # 将学号列转换为字符串格式
    data["学号"] = data["学号"].astype(str)
    
    # 选择所需列并添加表头
    data = data[columns]
    ws.append(columns)
    
    # 设置列宽
    for idx, col_width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = col_width
    
    # 写入数据
    for _, row in data.iterrows():
        ws.append(row.tolist())
    
    # 隐藏指定列
    for col in hidden_cols:
        col_idx = columns.index(col) + 1
        ws.column_dimensions[get_column_letter(col_idx)].hidden = True
    
    # 设置样式
    for i,row in enumerate(ws.iter_rows(min_row=1)):
        ws.row_dimensions[row[0].row].height = 27
        for cell in row:
            if i == 0:
                cell.font = 表头字体
            else:
                cell.font = 内容字体
            cell.alignment = 对齐方式
            cell.border = 边框样式

# 确认签字表样式设置
def style_confirm_sheet(ws,start_row,end_row,start_col,end_col,col_width=20,row_height=27):
    # 样式设置变量
    标题字体 = Font(size=18, bold=True)
    日期字体 = Font(size=14, bold=True)
    表头字体 = Font(size=12, bold=True)
    内容字体 = Font(size=12)    
    边框样式 = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))    
    对齐方式 = Alignment(horizontal="center", vertical="center")

    # 设置样式
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = 边框样式
            cell.alignment = 对齐方式
            if start_row == 1:
                cell.font = 标题字体
            elif start_row == 2:
                cell.font = 日期字体
            elif start_row == 3:
                cell.font = 表头字体
            else:
                cell.font = 内容字体

    # 设置列宽和行高
    for col_num in range(start_col, end_col + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = col_width
    for row_num in range(start_row, end_row + 1):
        ws.row_dimensions[row_num].height = row_height

# 学期汇总表样式设置
def style_summary_sheet(ws,row_height=27):
    # 样式设置变量
    表头字体 = Font(size=12, bold=True)
    内容字体 = Font(size=12)    
    边框样式 = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))    
    对齐方式 = Alignment(horizontal="center", vertical="center")

    # 设置样式
    for i,row in enumerate(ws.iter_rows(min_row=1)):
        ws.row_dimensions[row[0].row].height = row_height
        for cell in row:
            if i == 0:
                cell.font = 表头字体
            else:
                cell.font = 内容字体
            cell.alignment = 对齐方式
            cell.border = 边框样式

# 违规违纪名单写入及格式设置
def write_circular_sheet(ws, data, headers, col_widths, title, time_str, reason_str=None, type_col_name=None):

    # 样式设置变量
    对齐方式 = Alignment(horizontal='center', vertical='center')
    标题字体 = Font(size=16, bold=True)
    表头字体 = Font(size=12, bold=True)
    内容字体 = Font(size=12)
    边框样式 = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))

    # 写入标题
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1).value = title

    # 写入表头
    ws.append(headers)
    ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=6)

    # 设置列宽
    for idx, col_width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = col_width

    # 写入数据
    for index, (_, row) in enumerate(data.iterrows()):
        if type_col_name:
            if 10 <= row['旷课课时'] < 20:
                type_value = "警告"
                reason_str_l = reason_str[0]
            elif 20 <= row['旷课课时'] < 30:
                type_value = "严重警告"
                reason_str_l = reason_str[1]
            else:
                type_value = ""
            ws.append([index + 1, row['姓名'], time_str, row['班级'], reason_str_l, row['旷课课时'], type_value])
        else:
            ws.append([index + 1, row['姓名'], time_str, row['班级'], reason_str, row['旷课课时']])

    # 设置样式
    for i,row in enumerate(ws.iter_rows(min_row=1)):
        ws.row_dimensions[row[0].row].height = 27
        for cell in row:
            if i == 0:
                cell.font = 标题字体
            elif i == 1:
                cell.font = 表头字体
            else:
                cell.font = 内容字体
            cell.alignment = 对齐方式
            cell.border = 边框样式

# pdf 注册宋体
def register_chinese_font(font_name='SimSun', font_file=f'font/simsun.ttf'):
    # 判断是否为 PyInstaller 打包环境
    if hasattr(sys, '_MEIPASS'):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
    font_path = os.path.join(base_path, font_file)
    try:
        pdfmetrics.registerFont(TTFont(font_name, font_path))
    except Exception as e:
        print(f"Warning: {font_name} font not found, using fallback font")

# pdf 样式设置
def get_pdf_styles(cn_font='SimSun'):
    styles = getSampleStyleSheet()
    style_title = ParagraphStyle(
        'Title',
        parent=styles['Normal'],
        fontName=cn_font,
        fontSize=22,
        leading=24,
        alignment=TA_CENTER,
        spaceAfter=10*mm,
        textColor=colors.black
    )
    style_content = ParagraphStyle(
        'Content',
        parent=styles['Normal'],
        fontName=cn_font,
        fontSize=16,
        leading=20,
        alignment=TA_JUSTIFY,
        wordWrap='CJK',
        spaceAfter=8*mm
    )
    style_table_header = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName=cn_font,
        fontSize=12,
        alignment=TA_CENTER,
        textColor=colors.black,
        leading=14
    )
    style_table_cell = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName=cn_font,
        fontSize=12,
        alignment=TA_CENTER,
        leading=14
    )
    style_note = ParagraphStyle(
        'Note',
        parent=styles['Normal'],
        fontName=cn_font,
        fontSize=12,
        leading=16,
        alignment=TA_LEFT,
        spaceAfter=6*mm
    )
    style_sign = ParagraphStyle(
        'Sign',
        parent=styles['Normal'],
        fontName=cn_font,
        fontSize=16,
        leading=20,
        alignment=TA_LEFT,
        spaceAfter=6*mm
    )
    style_date = ParagraphStyle(
        'Date',
        parent=styles['Normal'],
        fontName=cn_font,
        fontSize=16,
        leading=20,
        alignment=TA_RIGHT
    )
    return {
        'title': style_title,
        'content': style_content,
        'table_header': style_table_header,
        'table_cell': style_table_cell,
        'note': style_note,
        'sign': style_sign,
        'date': style_date,
    }

# 压缩文件夹
def zip_files(dirs, output):
    zip_output = f"{output}.zip"
    with zipfile.ZipFile(zip_output, "w") as zipf:
        for dir in dirs:
            for root, _, files in os.walk(dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    if len(dirs) > 1:
                        arcname = os.path.join(os.path.basename(dir), os.path.relpath(file_path, start=dir))
                    else:
                        arcname = os.path.relpath(file_path, start=dir)
                    zipf.write(file_path, arcname)
    zipf.close()

# 下载完成后删除文件夹及文件
def delete_files_and_folders(paths):
    for path in paths:
        if os.path.exists(path):
            if os.path.isdir(path):
                shutil.rmtree(path)
            else:
                os.remove(path)
