import os
import sys

script_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, script_dir)

from tools import is_chinese, style_confirm_sheet, style_summary_sheet, write_circular_sheet

import re
import pandas as pd
from openpyxl import Workbook   
from openpyxl.utils.dataframe import dataframe_to_rows   
from openpyxl.utils import get_column_letter

def process_confirm_sheets(data, study_year, semester, start_year, start_month, 
                           start_day, end_year, end_month, end_day, output_folder_1, output_folder_2, class_list):
    
    data["姓名是否中文"] = data["姓名"].apply(is_chinese)
    data_cn = data[data["姓名是否中文"]]
    # 筛选班级
    pattern = "|".join(map(re.escape, class_list))
    unique_classes = [
        cls for cls in data_cn["班级"].dropna().unique() 
        if re.search(pattern, cls)
    ]
    
    # 确认签字表
    def create_confirm_sheet():

        # 设置样式并写入内容
        for class_name in unique_classes:
            class_data = data[data["班级"] == class_name]
            filtered_data = class_data[(class_data["旷课课时"] > 0)]
            filtered_data = filtered_data[["姓名", "旷课次数", "迟到次数", "早退次数", "旷课课时"]]
            filtered_data["确认签字"] = ""
            filtered_data = filtered_data.sort_values(by=["旷课课时", "旷课次数"], ascending=[False, False])

            wb = Workbook()
            ws = wb.active
            ws.merge_cells("A1:F1")
            ws["A1"] = f"{study_year}年度{class_name}上课啦扣分情况统计表"
            style_confirm_sheet(ws, 1, 1, 1, 6, col_width=30, row_height=30)

            ws.merge_cells("A2:F2")
            ws["A2"] = f"{start_year}年{start_month}月{start_day}日-{end_year}年{end_month}月{end_day}日"
            style_confirm_sheet(ws, 2, 2, 1, 6, col_width=30, row_height=25)

            header = ["姓名", "旷课次数", "迟到次数", "早退次数", "旷课课时", "确认签字"]
            ws.append(header)
            style_confirm_sheet(ws, 3, 3, 1, 6)

            for row in dataframe_to_rows(filtered_data, index=False, header=False):
                ws.append(row)

            style_confirm_sheet(ws, 4, ws.max_row, 1, 6)

            output_file = f"{output_folder_1}/{class_name}.xlsx"
            wb.save(output_file)
    
    # 学期汇总表
    def create_summary_sheet():

        # 筛选数据
        filtered_data = data_cn[(data_cn["旷课课时"] > 0)]
        filtered_data = filtered_data[["学号", "姓名", "班级", "旷课次数", "迟到次数", "早退次数", "旷课课时"]]
        filtered_data["学号"] = filtered_data["学号"].astype(str)
        filtered_data = filtered_data.sort_values(by=["旷课课时", "旷课次数"], ascending=[False, False])

        wb = Workbook()
        ws = wb.active

        colum_widths = [15, 11, 35, 10, 10, 10, 10]
        header = ["学号", "姓名", "班级", "旷课次数", "迟到次数", "早退次数", "旷课课时"]
        ws.append(header)

        for idx, col_width in enumerate(colum_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = col_width

        for row in dataframe_to_rows(filtered_data, index=False, header=False):
            ws.append(row)
        style_summary_sheet(ws)

        output_file = f"{output_folder_2}/{study_year}上课啦汇总.xlsx"
        wb.save(output_file)

    # 违规违纪名单
    def create_circular_sheet():
        wb = Workbook()
        header_common = ["序号", "姓名", "时间", "班级", "原因", ""]
        header_sheet2 = ["序号", "姓名", "时间", "班级", "原因", "", "类型"]

        title = f"{study_year}学年{semester}计算机科学与技术学院违规违纪名单"
        time_str = f"{study_year}学年{semester}"
        reason_1 = "旷课课时满5学时不足10学时"
        reason_2 = ("旷课课时满10学时不足20学时", "旷课课时满20学时不足30学时")
        reason_3 = "旷课课时满30学时不足40学时"
        reason_4 = "旷课课时40学时以上"

        # sheet2特殊一点
        column_widths_common = [5, 12, 35, 40, 31, 6]
        column_widths_sheet2 = [5, 12, 35, 40, 31, 6, 10]

        filtered_data = data_cn[(data_cn["旷课课时"] >= 5) & data_cn["班级"].isin(unique_classes)]
        filtered_data = filtered_data.sort_values(by=["旷课课时", "旷课次数"], ascending=[True, True])
        sheet1_data = filtered_data[(filtered_data['旷课课时'] >= 5) & (filtered_data['旷课课时'] < 10)]
        sheet2_data = filtered_data[(filtered_data['旷课课时'] >= 10) & (filtered_data['旷课课时'] < 30)]
        sheet3_data = filtered_data[(filtered_data['旷课课时'] >= 30) & (filtered_data['旷课课时'] < 40)]
        sheet4_data = filtered_data[filtered_data['旷课课时'] >= 40]

        sheet1 = wb.active
        sheet1.title = "sheet1"
        write_circular_sheet(sheet1, sheet1_data, header_common, column_widths_common, title, time_str, reason_1)

        sheet2 = wb.create_sheet(title="sheet2")
        write_circular_sheet(sheet2, sheet2_data, header_sheet2, column_widths_sheet2, title, time_str, reason_2, type_col_name="类型")

        sheet3 = wb.create_sheet(title="sheet3")
        write_circular_sheet(sheet3, sheet3_data, header_common, column_widths_common, title, time_str, reason_3)

        sheet4 = wb.create_sheet(title="sheet4")
        write_circular_sheet(sheet4, sheet4_data, header_common, column_widths_common, title, time_str, reason_4)

        output_file = f"{output_folder_2}/{study_year}上课啦违规违纪名单.xlsx"
        wb.save(output_file)

    create_confirm_sheet()
    create_summary_sheet()
    create_circular_sheet()
